import os
from openpyxl import load_workbook

def modify_cells_with_filename(filename):
    print(f"Processing file: {filename}")
    
    try:
        # Load the Excel workbook
        wb = load_workbook(filename)
    except Exception as e:
        print(f"Error loading workbook: {e}")
        return
    
    # Select the desired sheet
    sheet = wb.active
    
    # Get the base name of the file (excluding extension)
    base_name = os.path.splitext(os.path.basename(filename))[0]
    
    # Iterate through the range H2 to H29
    for row_number in range(2, 30):
        cell_value = sheet.cell(row=row_number, column=8).value  # Column H is the 8th column
        # Update cell value only if it's not already equal to the base name
        if cell_value != base_name:
            sheet.cell(row=row_number, column=8, value=base_name)
    
    # Save the modified workbook
    wb.save(filename)

# Specify the directory containing your Excel files
directory_path = r'C:\Users\kuang\Desktop\reporttt'

# Iterate through each file in the directory
for filename in os.listdir(directory_path):
    if filename.endswith('.xlsx'):
        file_path = os.path.join(directory_path, filename)
        modify_cells_with_filename(file_path)

print("Modification complete.")
